from datetime import UTC, date, datetime, time, timedelta
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4
from concurrent.futures import ThreadPoolExecutor
from time import perf_counter

import pytest
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from PIL import Image
import os

from sqlalchemy import create_engine, delete, func, select, text

from app.db.session import SessionLocal
from app.models.core import Client, Event, EventStaff, User
from app.models.enums import (
    EventStatus,
    LogbookAssignmentMode,
    LogbookAssignmentStatus,
    LogbookInstanceStatus,
    LogbookOperationalStage,
    LogbookTemplateStatus,
    LogbookVersionStatus,
    UserRole,
)
from app.models.logbook import (
    LogbookAssignment,
    LogbookContributionEvidence,
    LogbookImportBatch,
    LogbookInstance,
    LogbookInstanceItem,
    LogbookItemContribution,
    LogbookTemplate,
    LogbookTemplateVersion,
)
from app.services import logbook_contribution_service as contribution_service
from app.services import logbook_excel_service as excel_service
from app.services import logbook_service
from app.schemas.logbook_schema import MyAssignmentRead


def matrix(days=3):
    book = Workbook()
    sheet = book.active
    sheet.title = "Plan"
    sheet.cell(3, 2, "Actividad")
    for column in range(3, 3 + days):
        sheet.cell(3, column, datetime(2026, 2, 23) + timedelta(days=column - 3))
    for row, title in enumerate(("Oficinas", "Baños"), 4):
        sheet.cell(row, 2, title)
        for column in range(3, 3 + days):
            sheet.cell(row, column, "X")
    output = BytesIO()
    book.save(output)
    return output.getvalue()


@pytest.fixture()
def daily():
    db = SessionLocal()
    suffix = uuid4().hex[:8]
    client = Client(business_name=f"Daily {suffix}", contact_email=f"daily-{suffix}@example.com")
    other_client = Client(
        business_name=f"Daily other {suffix}",
        contact_email=f"daily-other-{suffix}@example.com",
    )
    db.add_all([client, other_client])
    db.flush()
    users = {}
    for name, role in (
        ("admin", UserRole.ADMIN),
        ("supervisor", UserRole.SUPERVISOR),
        ("a", UserRole.WORKER),
        ("b", UserRole.LOGISTICS_OPERATOR),
        ("c", UserRole.WORKER),
        ("outsider", UserRole.WORKER),
        ("client_owner", UserRole.CLIENT),
        ("client_external", UserRole.CLIENT),
    ):
        users[name] = User(
            full_name=f"User {name}",
            email=f"{name}-{suffix}@example.com",
            password_hash="x",
            role=role,
            is_active=True,
            client_id=(
                client.id
                if name == "client_owner"
                else other_client.id
                if name == "client_external"
                else None
            ),
        )
    db.add_all(users.values())
    db.flush()
    event = Event(
        client_id=client.id,
        name=f"Daily event {suffix}",
        start_date=datetime(2026, 2, 20),
        end_date=datetime(2026, 3, 31),
        status=EventStatus.PLANNING,
        created_by=users["admin"].id,
    )
    db.add(event)
    db.flush()
    db.add_all(
        [EventStaff(event_id=event.id, user_id=users[name].id) for name in ("supervisor", "a", "b", "c")]
    )
    template = LogbookTemplate(
        name=f"Daily template {suffix}",
        operational_stage=LogbookOperationalStage.OPERATION,
        status=LogbookTemplateStatus.ACTIVE,
        default_assignment_mode=LogbookAssignmentMode.SHARED,
        created_by=users["admin"].id,
    )
    db.add(template)
    db.flush()
    version = LogbookTemplateVersion(
        template_id=template.id,
        version_number=1,
        status=LogbookVersionStatus.PUBLISHED,
        created_by=users["admin"].id,
        published_by=users["admin"].id,
        published_at=datetime.now(UTC).replace(tzinfo=None),
    )
    db.add(version)
    db.commit()
    yield db, event, version, users
    db.rollback()
    db.execute(delete(LogbookInstance).where(LogbookInstance.event_id == event.id))
    db.execute(delete(LogbookImportBatch).where(LogbookImportBatch.event_id == event.id))
    db.execute(delete(EventStaff).where(EventStaff.event_id == event.id))
    db.delete(event)
    db.delete(template)
    db.execute(delete(User).where(User.id.in_([user.id for user in users.values()])))
    db.delete(client)
    db.delete(other_client)
    db.commit()
    db.close()


def config(version, users, digest):
    return SimpleNamespace(
        file_sha256=digest,
        template_version_id=version.id,
        participant_ids=[users["a"].id, users["b"].id],
        supervisor_id=users["supervisor"].id,
        opens_at_local=time(8),
        due_at_local=time(18),
        timezone="America/Santiago",
        client_visibility=False,
        base_name="Bitácora diaria",
    )


def test_preview_is_read_only_confirm_is_per_day_and_idempotent(daily):
    db, event, version, users = daily
    content = matrix(3)
    before = db.scalar(
        select(func.count(LogbookInstance.id)).where(LogbookInstance.event_id == event.id)
    )
    preview = excel_service.preview(db, event.id, content, "plan.xlsx", users["admin"])
    assert preview["instances_to_create"] == 3 and preview["scheduled_items_count"] == 6
    assert (
        db.scalar(
            select(func.count(LogbookInstance.id)).where(LogbookInstance.event_id == event.id)
        )
        == before
    )
    result = excel_service.import_xlsx(
        db,
        event.id,
        content,
        "plan.xlsx",
        config(version, users, preview["file_sha256"]),
        users["admin"],
    )
    assert result["instances_created"] == 3
    assert (
        db.scalar(
            select(func.count(LogbookInstanceItem.id))
            .join(LogbookInstance)
            .where(LogbookInstance.event_id == event.id)
        )
        == 6
    )
    with pytest.raises(HTTPException) as duplicate:
        excel_service.import_xlsx(
            db,
            event.id,
            content,
            "plan.xlsx",
            config(version, users, preview["file_sha256"]),
            users["admin"],
        )
    assert duplicate.value.status_code == 409


def test_worker_assignment_list_includes_instance_schedule(daily):
    db, event, version, users = daily
    content = matrix(1)
    preview = excel_service.preview(db, event.id, content, "worker-list.xlsx", users["admin"])
    imported = excel_service.import_xlsx(
        db, event.id, content, "worker-list.xlsx",
        config(version, users, preview["file_sha256"]), users["admin"],
    )
    rows = logbook_service.my_assignments(db, users["a"])
    payload = MyAssignmentRead.model_validate(rows[0])
    assert payload.logbook_instance_id == imported["instance_ids"][0]
    assert payload.instance.name == "Bitácora diaria · 23/02/2026"
    assert payload.instance.occurrence_date == date(2026, 2, 23)
    assert payload.instance.opens_at is not None
    assert payload.instance.due_at is not None


def test_direct_logbook_assignment_allows_detail_without_broad_event_access(daily):
    db, event, version, users = daily
    event.status = EventStatus.QUOTE
    db.commit()
    content = matrix(1)
    preview = excel_service.preview(db, event.id, content, "direct-access.xlsx", users["admin"])
    imported = excel_service.import_xlsx(
        db, event.id, content, "direct-access.xlsx",
        config(version, users, preview["file_sha256"]), users["admin"],
    )
    detail = logbook_service.get_instance_detail(db, imported["instance_ids"][0], users["a"])
    assert detail["id"] == imported["instance_ids"][0]
    assert [assignment["user_id"] for assignment in detail["assignments"]] == [users["a"].id]
    with pytest.raises(HTTPException) as outsider:
        logbook_service.get_instance_detail(db, imported["instance_ids"][0], users["outsider"])
    assert outsider.value.status_code == 403


def test_generated_template_has_real_dates_protection_and_round_trips(daily):
    db, event, _, users = daily
    content, filename = excel_service.generate_template(
        db, event.id, date(2026, 2, 23), date(2026, 2, 25), users["admin"]
    )
    assert filename.endswith("-2026-02-23-2026-02-25.xlsx")
    book = load_workbook(BytesIO(content))
    assert book.sheetnames == ["Planificación", "Instrucciones"]
    sheet = book["Planificación"]
    assert sheet["A3"].value == "Actividad"
    assert [sheet.cell(3, column).value for column in range(2, 5)] == [
        datetime(2026, 2, 23), datetime(2026, 2, 24), datetime(2026, 2, 25)
    ]
    assert all(sheet.cell(3, column).number_format == "dd/mm/yyyy" for column in range(2, 5))
    assert sheet.protection.sheet is True
    assert sheet["A3"].protection.locked is True
    assert sheet["A4"].protection.locked is False
    assert sheet["B4"].protection.locked is False
    assert sheet.freeze_panes == "B4"
    assert len(sheet.data_validations.dataValidation) == 1
    assert "Evento:" in book["Instrucciones"]["A8"].value

    sheet["A4"] = "Limpieza oficinas"
    sheet["B4"] = "X"
    sheet["D4"] = "x"
    completed = BytesIO()
    book.save(completed)
    started = perf_counter()
    parsed = excel_service.parse_xlsx(completed.getvalue(), filename, event)
    assert perf_counter() - started < 5
    assert parsed["activities_count"] == 1
    assert parsed["dates_count"] == 3
    assert parsed["scheduled_items_count"] == 2
    assert parsed["instances_to_create"] == 2
    assert parsed["errors"] == []


def test_generated_template_rejects_invalid_ranges_and_unauthorized_users(daily):
    db, event, _, users = daily
    invalid = [
        (date(2026, 2, 25), date(2026, 2, 23)),
        (date(2026, 2, 19), date(2026, 2, 23)),
        (date(2026, 2, 23), date(2026, 4, 1)),
    ]
    for start, end in invalid:
        with pytest.raises(HTTPException) as error:
            excel_service.generate_template(db, event.id, start, end, users["admin"])
        assert error.value.status_code == 422
    for name in ("outsider", "client_owner", "client_external"):
        with pytest.raises(HTTPException) as denied:
            excel_service.generate_template(
                db, event.id, date(2026, 2, 23), date(2026, 2, 25), users[name]
            )
        assert denied.value.status_code == 403


def test_sha_mismatch_and_atomic_rollback(daily, monkeypatch):
    db, event, version, users = daily
    content = matrix(10)
    preview = excel_service.preview(db, event.id, content, "plan.xlsx", users["admin"])
    wrong = config(version, users, "0" * 64)
    with pytest.raises(HTTPException) as mismatch:
        excel_service.import_xlsx(db, event.id, content, "plan.xlsx", wrong, users["admin"])
    assert mismatch.value.status_code == 409
    original, calls = excel_service.LogbookInstanceItem, {"count": 0}

    def fail_late(*args, **kwargs):
        calls["count"] += 1
        if calls["count"] == 15:
            raise RuntimeError("induced atomicity failure")
        return original(*args, **kwargs)

    monkeypatch.setattr(excel_service, "LogbookInstanceItem", fail_late)
    with pytest.raises(RuntimeError, match="induced"):
        excel_service.import_xlsx(
            db,
            event.id,
            content,
            "plan.xlsx",
            config(version, users, preview["file_sha256"]),
            users["admin"],
        )
    assert (
        db.scalar(
            select(func.count(LogbookImportBatch.id)).where(LogbookImportBatch.event_id == event.id)
        )
        == 0
    )
    assert (
        db.scalar(
            select(func.count(LogbookInstance.id)).where(LogbookInstance.event_id == event.id)
        )
        == 0
    )
    assert (
        db.scalar(
            select(func.count(LogbookInstanceItem.id))
            .join(LogbookInstance)
            .where(LogbookInstance.event_id == event.id)
        )
        == 0
    )
    assert (
        db.scalar(
            select(func.count(LogbookAssignment.id))
            .join(LogbookInstance)
            .where(LogbookInstance.event_id == event.id)
        )
        == 0
    )
    assert (
        db.scalar(
            select(func.count(LogbookItemContribution.id))
            .join(LogbookInstanceItem)
            .join(LogbookInstance)
            .where(LogbookInstance.event_id == event.id)
        )
        == 0
    )


def test_two_workers_contribute_without_overwrite_and_evidence_is_private(
    daily, tmp_path, monkeypatch
):
    db, event, version, users = daily
    content = matrix(1)
    preview = excel_service.preview(db, event.id, content, "plan.xlsx", users["admin"])
    result = excel_service.import_xlsx(
        db,
        event.id,
        content,
        "plan.xlsx",
        config(version, users, preview["file_sha256"]),
        users["admin"],
    )
    instance_id = result["instance_ids"][0]
    item = db.scalar(
        select(LogbookInstanceItem).where(LogbookInstanceItem.instance_id == instance_id)
    )
    a = contribution_service.save(
        db, item.id, SimpleNamespace(description="Limpié escritorios", version=None), users["a"]
    )
    b = contribution_service.save(
        db, item.id, SimpleNamespace(description="Limpié pisos", version=None), users["b"]
    )
    assert a.id != b.id and {a.description, b.description} == {"Limpié escritorios", "Limpié pisos"}
    with pytest.raises(HTTPException) as conflict:
        contribution_service.save(
            db, item.id, SimpleNamespace(description="viejo", version=999), users["a"]
        )
    assert conflict.value.status_code == 409
    with pytest.raises(HTTPException) as outsider:
        contribution_service.save(
            db, item.id, SimpleNamespace(description="intruso", version=None), users["outsider"]
        )
    assert outsider.value.status_code == 403
    image = BytesIO()
    Image.new("RGB", (10, 10), "green").save(image, "JPEG")
    monkeypatch.setattr(
        "app.services.file_storage_service.settings.local_private_storage_root", str(tmp_path)
    )
    for setting_name in (
        "cloudflare_r2_bucket", "cloudflare_r2_account_id",
        "cloudflare_r2_access_key_id", "cloudflare_r2_secret_access_key",
    ):
        monkeypatch.setattr(
            f"app.services.file_storage_service.settings.{setting_name}", None
        )
    upload = UploadFile(
        filename="proof.jpg", file=BytesIO(image.getvalue()), headers={"content-type": "image/jpeg"}
    )
    evidence = contribution_service.upload_evidence(db, a.id, upload, users["a"])
    assert evidence.storage_key and evidence.assignment_id == a.assignment_id
    event.status = EventStatus.QUOTE
    db.commit()
    author_access = contribution_service.evidence_access(db, evidence.id, users["a"])
    admin_access = contribution_service.evidence_access(db, evidence.id, users["admin"])
    assert author_access["expires_in"] == admin_access["expires_in"] == 300
    author_token = author_access["url"].split("token=", 1)[1]
    stored_content, stored_type, stored_name = contribution_service.evidence_content(
        db, evidence.id, author_token
    )
    assert stored_content == image.getvalue()
    assert stored_type == "image/jpeg" and stored_name == "proof.jpg"
    with pytest.raises(HTTPException) as other_delete:
        contribution_service.delete_evidence(db, evidence.id, users["b"])
    assert other_delete.value.status_code == 403
    with pytest.raises(HTTPException) as other_read:
        contribution_service.evidence_access(db, evidence.id, users["b"])
    assert other_read.value.status_code == 403
    event.status = EventStatus.PLANNING
    db.commit()
    access = contribution_service.evidence_access(db, evidence.id, users["supervisor"])
    assert "storage_key" not in access and access["expires_in"] == 300
    contribution_service.delete_evidence(db, evidence.id, users["a"])
    assert db.get(LogbookContributionEvidence, evidence.id).deleted_at is not None


def test_multiple_timestamped_records_and_role_permissions(daily):
    db, event, version, users = daily
    content = matrix(1)
    preview = excel_service.preview(db, event.id, content, "timeline.xlsx", users["admin"])
    result = excel_service.import_xlsx(
        db, event.id, content, "timeline.xlsx",
        config(version, users, preview["file_sha256"]), users["admin"],
    )
    instance_id = result["instance_ids"][0]
    item = db.scalar(select(LogbookInstanceItem).where(
        LogbookInstanceItem.instance_id == instance_id
    ))
    first = contribution_service.create(
        db, item.id, SimpleNamespace(description="Initial cleaning", version=None), users["a"]
    )
    second = contribution_service.create(
        db, item.id, SimpleNamespace(description="Afternoon cleaning", version=None), users["a"]
    )
    logistics = contribution_service.create(
        db, item.id, SimpleNamespace(description="Material removal", version=None), users["b"]
    )
    assert len({first.id, second.id, logistics.id}) == 3
    assert contribution_service.metrics(db, instance_id, users["supervisor"])[
        "contributions_count"
    ] == 3
    imported_metrics = logbook_service.calculate_instance_metrics(db, item.instance)
    assert imported_metrics["completion_percentage"] == 50.0
    assert imported_metrics["participation_percentage"] == 100.0

    updated = contribution_service.update(
        db, first.id,
        SimpleNamespace(description="Initial cleaning corrected", version=first.version),
        users["a"],
    )
    assert updated.version == 2 and updated.description == "Initial cleaning corrected"
    with pytest.raises(HTTPException) as other_update:
        contribution_service.update(
            db, updated.id,
            SimpleNamespace(description="Unauthorized", version=updated.version), users["b"],
        )
    assert other_update.value.status_code == 403
    with pytest.raises(HTTPException) as stale_update:
        contribution_service.update(
            db, updated.id,
            SimpleNamespace(description="Stale", version=1), users["a"],
        )
    assert stale_update.value.status_code == 409

    contribution_service.remove(db, second.id, second.version, users["a"])
    for viewer in (users["admin"], users["supervisor"], users["a"], users["b"]):
        visible = contribution_service.list_items(db, instance_id, viewer)
        assert [entry["description"] for entry in visible[0]["contributions"]] == [
            "Initial cleaning corrected", "Material removal",
        ]
    for forbidden in (users["outsider"], users["client_owner"], users["client_external"]):
        with pytest.raises(HTTPException) as denied:
            contribution_service.list_items(db, instance_id, forbidden)
        assert denied.value.status_code == 403


def test_database_rejects_cross_instance_assignment(daily):
    db, event, version, users = daily
    content = matrix(2)
    preview = excel_service.preview(db, event.id, content, "plan.xlsx", users["admin"])
    result = excel_service.import_xlsx(
        db,
        event.id,
        content,
        "plan.xlsx",
        config(version, users, preview["file_sha256"]),
        users["admin"],
    )
    first, second = result["instance_ids"]
    item = db.scalar(select(LogbookInstanceItem).where(LogbookInstanceItem.instance_id == first))
    assignment = db.scalar(
        select(LogbookAssignment).where(LogbookAssignment.logbook_instance_id == second)
    )
    db.add(
        LogbookItemContribution(
            instance_item_id=item.id,
            assignment_id=assignment.id,
            instance_id=first,
            author_id=assignment.user_id,
            description="crossed",
        )
    )
    with pytest.raises(Exception):
        db.flush()
    db.rollback()


def test_concurrent_duplicate_import_creates_one_batch(daily):
    db, event, version, users = daily
    content = matrix(4)
    preview = excel_service.preview(db, event.id, content, "concurrent.xlsx", users["admin"])
    event_id, version_id = event.id, version.id
    participant_ids = [users["a"].id, users["b"].id]
    supervisor_id, admin_id = users["supervisor"].id, users["admin"].id

    def attempt():
        session = SessionLocal()
        try:
            actor = session.get(User, admin_id)
            payload = SimpleNamespace(
                file_sha256=preview["file_sha256"],
                template_version_id=version_id,
                participant_ids=participant_ids,
                supervisor_id=supervisor_id,
                opens_at_local=time(8),
                due_at_local=time(18),
                timezone="America/Santiago",
                client_visibility=False,
                base_name="Concurrent",
            )
            return excel_service.import_xlsx(
                session, event_id, content, "concurrent.xlsx", payload, actor
            )
        except HTTPException as exc:
            return exc.status_code
        finally:
            session.close()

    with ThreadPoolExecutor(max_workers=2) as pool:
        outcomes = list(pool.map(lambda _: attempt(), range(2)))
    assert sum(isinstance(value, dict) for value in outcomes) == 1
    assert 409 in outcomes
    db.expire_all()
    assert (
        db.scalar(
            select(func.count(LogbookImportBatch.id)).where(LogbookImportBatch.event_id == event_id)
        )
        == 1
    )
    assert (
        db.scalar(
            select(func.count(LogbookInstance.id)).where(LogbookInstance.event_id == event_id)
        )
        == 4
    )


def test_bulk_participants_preview_scopes_and_preserves_history(daily):
    db, event, version, users = daily
    content = matrix(3)
    preview = excel_service.preview(db, event.id, content, "bulk.xlsx", users["admin"])
    imported = excel_service.import_xlsx(
        db, event.id, content, "bulk.xlsx",
        config(version, users, preview["file_sha256"]), users["admin"],
    )
    batch_id = imported["batch_id"]
    instances = list(db.scalars(select(LogbookInstance).where(
        LogbookInstance.import_batch_id == batch_id
    ).order_by(LogbookInstance.occurrence_date)).all())
    instances[-1].status = LogbookInstanceStatus.SCHEDULED
    db.commit()

    add_payload = SimpleNamespace(
        operation="ADD", scope="ALL", participant_ids=[users["c"].id], dates=[]
    )
    impact = excel_service.bulk_participants(db, batch_id, add_payload, users["admin"], apply=False)
    assert impact == {
        "batch_id": batch_id, "operation": "ADD", "scope": "ALL",
        "instances_matched": 3, "assignments_to_add": 3, "assignments_to_remove": 0,
        "assignments_preserved": 6, "historical_assignments_preserved": 0,
        "participant_ids": [users["c"].id], "applied": False,
    }
    assert db.scalar(select(func.count(LogbookAssignment.id)).join(LogbookInstance).where(
        LogbookInstance.import_batch_id == batch_id
    )) == 6
    applied = excel_service.bulk_participants(db, batch_id, add_payload, users["admin"], apply=True)
    assert applied["applied"] is True
    assert db.scalar(select(func.count(LogbookAssignment.id)).join(LogbookInstance).where(
        LogbookInstance.import_batch_id == batch_id
    )) == 9

    first_item = db.scalar(select(LogbookInstanceItem).where(
        LogbookInstanceItem.instance_id == instances[0].id
    ))
    contribution_service.save(
        db, first_item.id, SimpleNamespace(description="Histórico", version=None), users["a"]
    )
    remove_payload = SimpleNamespace(
        operation="REMOVE", scope="ALL", participant_ids=[users["a"].id], dates=[]
    )
    removal = excel_service.bulk_participants(db, batch_id, remove_payload, users["admin"], apply=True)
    assert removal["assignments_to_remove"] == 3
    assert removal["historical_assignments_preserved"] == 1
    historical_assignment = db.scalar(select(LogbookAssignment).where(
        LogbookAssignment.logbook_instance_id == instances[0].id,
        LogbookAssignment.user_id == users["a"].id,
    ))
    assert historical_assignment.status == LogbookAssignmentStatus.CANCELLED
    assert db.scalar(select(func.count(LogbookItemContribution.id)).where(
        LogbookItemContribution.assignment_id == historical_assignment.id
    )) == 1

    future_payload = SimpleNamespace(
        operation="REPLACE", scope="FUTURE", participant_ids=[users["b"].id], dates=[]
    )
    future = excel_service.bulk_participants(db, batch_id, future_payload, users["admin"], apply=True)
    assert future["instances_matched"] == 1
    active_future = set(db.scalars(select(LogbookAssignment.user_id).where(
        LogbookAssignment.logbook_instance_id == instances[-1].id,
        LogbookAssignment.status != LogbookAssignmentStatus.CANCELLED,
    )).all())
    assert active_future == {users["b"].id}


def test_bulk_participants_validates_scope_staff_and_permissions(daily):
    db, event, version, users = daily
    content = matrix(1)
    preview = excel_service.preview(db, event.id, content, "bulk-guard.xlsx", users["admin"])
    imported = excel_service.import_xlsx(
        db, event.id, content, "bulk-guard.xlsx",
        config(version, users, preview["file_sha256"]), users["admin"],
    )
    batch_id = imported["batch_id"]
    with pytest.raises(HTTPException) as outsider:
        excel_service.bulk_participants(
            db, batch_id,
            SimpleNamespace(operation="ADD", scope="ALL", participant_ids=[users["c"].id], dates=[]),
            users["outsider"], apply=False,
        )
    assert outsider.value.status_code == 403
    with pytest.raises(HTTPException) as client:
        excel_service.bulk_participants(
            db, batch_id,
            SimpleNamespace(operation="ADD", scope="ALL", participant_ids=[users["c"].id], dates=[]),
            users["client_owner"], apply=False,
        )
    assert client.value.status_code == 403
    with pytest.raises(HTTPException) as invalid_staff:
        excel_service.bulk_participants(
            db, batch_id,
            SimpleNamespace(operation="ADD", scope="ALL", participant_ids=[users["outsider"].id], dates=[]),
            users["admin"], apply=False,
        )
    assert invalid_staff.value.status_code == 422
    with pytest.raises(HTTPException) as empty:
        excel_service.bulk_participants(
            db, batch_id,
            SimpleNamespace(operation="REMOVE", scope="ALL", participant_ids=[users["a"].id,users["b"].id], dates=[]),
            users["admin"], apply=True,
        )
    assert empty.value.status_code == 422
    assert db.scalar(select(func.count(LogbookAssignment.id)).join(LogbookInstance).where(
        LogbookInstance.import_batch_id == batch_id
    )) == 2


def test_new_tables_rls_role_matrix(daily):
    db, event, version, users = daily
    content = matrix(1)
    preview = excel_service.preview(db, event.id, content, "rls.xlsx", users["admin"])
    result = excel_service.import_xlsx(
        db,
        event.id,
        content,
        "rls.xlsx",
        config(version, users, preview["file_sha256"]),
        users["admin"],
    )
    instance_id = result["instance_ids"][0]
    item = db.scalar(
        select(LogbookInstanceItem).where(LogbookInstanceItem.instance_id == instance_id)
    )
    contribution_a = contribution_service.save(
        db, item.id, SimpleNamespace(description="A", version=None), users["a"]
    )
    contribution_service.save(
        db, item.id, SimpleNamespace(description="B", version=None), users["b"]
    )
    evidence = LogbookContributionEvidence(
        contribution_id=contribution_a.id,
        event_id=event.id,
        instance_id=instance_id,
        instance_item_id=item.id,
        assignment_id=contribution_a.assignment_id,
        uploaded_by=users["a"].id,
        mime_type="image/jpeg",
        file_size=4,
        original_filename="private.jpg",
        storage_key="private/logbook-contributions/private.jpg",
    )
    db.add(evidence)
    db.commit()
    runtime = create_engine(os.environ["RLS_DATABASE_URL"])

    def counts(user, role):
        with runtime.begin() as connection:
            connection.execute(
                text("select set_config('app.current_user_id', :id, true)"), {"id": str(user.id)}
            )
            connection.execute(
                text("select set_config('app.current_role', :role, true)"), {"role": role.value}
            )
            connection.execute(
                text("select set_config('app.current_client_id', :id, true)"),
                {"id": str(user.client_id) if user.client_id else ""},
            )
            return (
                connection.scalar(
                    text("select count(*) from logbook_instance_items where instance_id=:id"),
                    {"id": instance_id},
                ),
                connection.scalar(
                    text("select count(*) from logbook_item_contributions where instance_id=:id"),
                    {"id": instance_id},
                ),
                connection.scalar(
                    text("select count(*) from logbook_import_batches where event_id=:id"),
                    {"id": event.id},
                ),
                connection.scalar(
                    text(
                        "select count(*) from logbook_contribution_evidences "
                        "where event_id=:id"
                    ),
                    {"id": event.id},
                ),
            )

    assert counts(users["admin"], UserRole.ADMIN) == (2, 2, 1, 1)
    assert counts(users["supervisor"], UserRole.SUPERVISOR) == (2, 2, 1, 1)
    assert counts(users["a"], UserRole.WORKER) == (2, 2, 0, 1)
    assert counts(users["outsider"], UserRole.WORKER) == (0, 0, 0, 0)
    assert counts(users["outsider"], UserRole.LOGISTICS_OPERATOR) == (0, 0, 0, 0)
    assert counts(users["client_owner"], UserRole.CLIENT) == (0, 0, 0, 0)
    assert counts(users["client_external"], UserRole.CLIENT) == (0, 0, 0, 0)
    for client_name in ("client_owner", "client_external"):
        client_user = users[client_name]
        with pytest.raises(HTTPException) as preview_denied:
            excel_service.preview(db, event.id, content, "rls.xlsx", client_user)
        assert preview_denied.value.status_code == 403
        with pytest.raises(HTTPException) as contribution_denied:
            contribution_service.save(
                db, item.id, SimpleNamespace(description="private", version=None), client_user
            )
        assert contribution_denied.value.status_code == 403
        with pytest.raises(HTTPException) as evidence_denied:
            contribution_service.evidence_access(db, evidence.id, client_user)
        assert evidence_denied.value.status_code == 403
    runtime.dispose()
