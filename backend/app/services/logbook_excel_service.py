from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import PurePath
from uuid import UUID
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.datetime import from_excel
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.models.core import Event, EventStaff, User
from app.models.enums import (
    LogbookAssignmentMode, LogbookAssignmentStatus, LogbookInstanceStatus,
    LogbookVersionStatus, UserRole,
)
from app.models.logbook import (
    LogbookAssignment, LogbookImportBatch, LogbookInstance, LogbookInstanceItem,
    LogbookItemContribution, LogbookTemplateVersion,
)
from app.models.enums import LogbookTemplateStatus
from app.services.logbook_service import audit
from app.core.permissions import can_manage_event

MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_ROWS = 1000
MAX_DATES = 366
MAX_CELLS = 200_000
MAX_TEMPLATE_ACTIVITY_ROWS = 997
HEADER = "actividad"
MONTHS = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
          "jul": 7, "ago": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dic": 12}


@dataclass(frozen=True)
class ParsedActivity:
    source_row: int
    title: str


def _issue(code: str, message: str, *, row=None, column=None, value=None) -> dict:
    return {k: v for k, v in {"code": code, "message": message, "row": row,
                              "column": column, "value": value}.items() if v is not None}


def _infer_date(value, event: Event, epoch) -> tuple[date | None, str | None]:
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch).date(), None
        except (TypeError, ValueError, OverflowError):
            return None, "Fecha Excel inválida"
    text = str(value or "").strip().lower().replace(".", "")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            pass
    match = re.fullmatch(r"(\d{1,2})[-/ ]([a-záéíóú]{3,10})", text)
    if not match:
        return None, "Encabezado de fecha no reconocido"
    month = MONTHS.get(match.group(2)[:4].rstrip("t")) or MONTHS.get(match.group(2)[:3])
    if not month:
        return None, "Mes no reconocido"
    start, end = event.start_date.date(), event.end_date.date()
    candidates = []
    for year in range(start.year, end.year + 1):
        try:
            candidate = date(year, month, int(match.group(1)))
        except ValueError:
            return None, "Fecha inválida"
        if start <= candidate <= end:
            candidates.append(candidate)
    if len(candidates) != 1:
        return None, "El año no puede inferirse inequívocamente desde el evento"
    return candidates[0], None


def parse_xlsx(content: bytes, filename: str, event: Event) -> dict:
    errors, warnings = [], []
    safe_name = PurePath(filename or "planificacion.xlsx").name[:255]
    digest = hashlib.sha256(content).hexdigest()
    if not safe_name.lower().endswith(".xlsx"):
        return {"filename": safe_name, "file_sha256": digest, "warnings": [],
                "errors": [_issue("invalid_extension", "El archivo debe ser .xlsx")]}
    if len(content) > MAX_FILE_SIZE:
        return {"filename": safe_name, "file_sha256": digest, "warnings": [],
                "errors": [_issue("file_too_large", "El archivo excede 5 MB")]}
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except Exception:
        return {"filename": safe_name, "file_sha256": digest, "warnings": [],
                "errors": [_issue("corrupt_workbook", "El libro no es un XLSX válido")]}
    try:
        if not workbook.sheetnames:
            raise ValueError("empty")
        located = None
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(max_row=min(sheet.max_row, MAX_ROWS)):
                for cell in row:
                    if str(cell.value or "").strip().casefold() == HEADER:
                        located = sheet, cell.row, cell.column
                        break
                if located:
                    break
            if located:
                break
        if not located:
            errors.append(_issue("activity_header_missing", "No se encontró el encabezado Actividad"))
            return _result(safe_name, digest, None, [], [], errors, warnings)
        sheet, header_row, activity_col = located
        if sheet.max_row > MAX_ROWS or sheet.max_column - activity_col > MAX_DATES or sheet.max_row * sheet.max_column > MAX_CELLS:
            errors.append(_issue("limits_exceeded", "La hoja excede los límites permitidos"))
            return _result(safe_name, digest, sheet.title, [], [], errors, warnings)
        rows = list(sheet.iter_rows(
            min_row=header_row, max_row=sheet.max_row,
            min_col=activity_col, max_col=sheet.max_column, values_only=True,
        ))
        header_values = rows[0]
        dates, seen_dates = [], {}
        for col in range(activity_col + 1, sheet.max_column + 1):
            raw = header_values[col - activity_col]
            if raw is None or str(raw).strip() == "":
                continue
            parsed, reason = _infer_date(raw, event, workbook.epoch)
            if reason:
                errors.append(_issue("invalid_date", reason, row=header_row, column=col, value=str(raw)[:100]))
                continue
            if parsed in seen_dates:
                errors.append(_issue("duplicate_date", "Fecha duplicada", row=header_row, column=col, value=parsed.isoformat()))
            else:
                seen_dates[parsed] = col
                dates.append((parsed, col))
            if not (event.start_date.date() <= parsed <= event.end_date.date()):
                errors.append(_issue("date_outside_event", "Fecha fuera del evento", row=header_row, column=col, value=parsed.isoformat()))
        activities, seen_titles, days = [], {}, {day: [] for day, _ in dates}
        for row, row_values in enumerate(rows[1:], header_row + 1):
            title_raw = row_values[0]
            values = [
                (day, col, row_values[col - activity_col])
                for day, col in dates
            ]
            marked = any(str(value or "").strip().casefold() == "x" for _, _, value in values)
            title = str(title_raw or "").strip()
            if not title:
                if marked:
                    errors.append(_issue("missing_activity", "Fila programada sin actividad", row=row, column=activity_col))
                continue
            key = title.casefold()
            if key in seen_titles:
                warnings.append(_issue("duplicate_activity", f"Actividad duplicada; también aparece en fila {seen_titles[key]}", row=row, column=activity_col, value=title))
            else:
                seen_titles[key] = row
            activity = ParsedActivity(row, title[:180])
            activities.append(activity)
            for day, col, value in values:
                normalized = str(value or "").strip()
                if not normalized:
                    continue
                if normalized.casefold() == "x":
                    days[day].append(activity)
                else:
                    warnings.append(_issue("unexpected_value", "Se esperaba X o una celda vacía", row=row, column=col, value=normalized[:100]))
        scheduled = sum(map(len, days.values()))
        if not scheduled:
            errors.append(_issue("nothing_scheduled", "No hay actividades programadas"))
        for day, items in days.items():
            if not items:
                warnings.append(_issue("empty_date", "La fecha no contiene actividades", value=day.isoformat()))
        return _result(safe_name, digest, sheet.title, activities, days, errors, warnings)
    finally:
        workbook.close()


def _result(filename, digest, sheet, activities, days, errors, warnings):
    dates = sorted(days) if isinstance(days, dict) else []
    scheduled = sum(len(value) for value in days.values()) if isinstance(days, dict) else 0
    return {"filename": filename, "file_sha256": digest, "sheet_name": sheet,
            "activities_count": len(activities), "dates_count": len(dates),
            "scheduled_items_count": scheduled,
            "instances_to_create": sum(bool(days[d]) for d in dates),
            "date_range": {"from": dates[0].isoformat(), "to": dates[-1].isoformat()} if dates else None,
            "warnings": warnings, "errors": errors,
            "days": [{"date": d.isoformat(), "activities": [{"source_row": a.source_row, "title": a.title} for a in days[d]]} for d in dates]}


def preview(db: Session, event_id: UUID, content: bytes, filename: str, current):
    if not can_manage_event(current, event_id, db):
        raise HTTPException(403, "Insufficient role")
    event = db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    return parse_xlsx(content, filename, event)


def generate_template(db: Session, event_id: UUID, start: date, end: date, current):
    if not can_manage_event(current, event_id, db):
        raise HTTPException(403, "Insufficient role")
    event = db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    if start > end:
        raise HTTPException(422, "La fecha inicial debe ser anterior o igual a la fecha final")
    event_start, event_end = event.start_date.date(), event.end_date.date()
    if start < event_start or end > event_end:
        raise HTTPException(
            422,
            f"El rango debe estar dentro del evento ({event_start:%d/%m/%Y}–{event_end:%d/%m/%Y})",
        )
    day_count = (end - start).days + 1
    if day_count > MAX_DATES:
        raise HTTPException(422, f"El rango no puede superar {MAX_DATES} días")
    dates = [start + timedelta(days=offset) for offset in range(day_count)]
    activity_rows = min(MAX_TEMPLATE_ACTIVITY_ROWS, (MAX_CELLS // (day_count + 1)) - 3)
    if activity_rows < 1:
        raise HTTPException(422, "El rango genera una plantilla demasiado grande")

    book = Workbook()
    sheet = book.active
    sheet.title = "Planificación"
    sheet["A1"] = event.name
    sheet["A2"] = f"Rango: {start:%d/%m/%Y} al {end:%d/%m/%Y}"
    sheet.cell(3, 1, "Actividad")
    for index, current_date in enumerate(dates, 2):
        cell = sheet.cell(3, index, current_date)
        cell.number_format = "dd/mm/yyyy"
    header_fill = PatternFill("solid", fgColor="166534")
    for cell in sheet[3][:day_count + 1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet["A1"].font = Font(size=14, bold=True, color="166534")
    sheet["A3"].comment = Comment("Escribe una actividad por fila.", "EcoEvent 360")
    sheet.freeze_panes = "B4"
    sheet.column_dimensions["A"].width = 42
    for column in range(2, day_count + 2):
        sheet.column_dimensions[sheet.cell(3, column).column_letter].width = 13
    last_row = 3 + activity_rows
    validation = DataValidation(type="list", formula1='"X,x"', allow_blank=True)
    validation.promptTitle = "Programar actividad"
    validation.prompt = "Selecciona X cuando la actividad corresponda a esta fecha."
    validation.error = "Solo se permite X o una celda vacía."
    validation.errorTitle = "Valor no válido"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(f"B4:{sheet.cell(3, day_count + 1).column_letter}{last_row}")
    for row in sheet.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=day_count + 1):
        for cell in row:
            cell.protection = Protection(locked=False)
    sheet.protection.sheet = True
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True

    instructions = book.create_sheet("Instrucciones")
    instructions.column_dimensions["A"].width = 105
    lines = [
        "Plantilla oficial EcoEvent 360",
        "1. Trabaja únicamente en la hoja Planificación.",
        "2. Escribe una actividad por fila en la columna Actividad.",
        "3. Marca X en cada fecha donde corresponda ejecutar la actividad.",
        "4. Deja la celda vacía cuando la actividad no corresponda.",
        "5. No cambies, borres ni agregues columnas de fechas.",
        "6. No repitas nombres de actividades.",
        f"Evento: {event.name}",
        f"Rango autorizado: {start:%d/%m/%Y} al {end:%d/%m/%Y} ({day_count} días).",
        f"Capacidad de esta plantilla: {activity_rows} actividades.",
    ]
    for row, line in enumerate(lines, 1):
        instructions.cell(row, 1, line)
    instructions["A1"].font = Font(size=16, bold=True, color="166534")
    instructions.protection.sheet = True
    book.active = 0
    output = io.BytesIO()
    book.save(output)
    safe_event = re.sub(r"[^A-Za-z0-9_-]+", "-", event.name).strip("-")[:80] or "evento"
    filename = f"plantilla-bitacoras-{safe_event}-{start.isoformat()}-{end.isoformat()}.xlsx"
    audit(db, current, "LOGBOOK_XLSX_TEMPLATE_DOWNLOADED", "Event", event.id,
          event_id=event.id, metadata={"start_date": start, "end_date": end, "days": day_count})
    db.commit()
    return output.getvalue(), filename


def import_xlsx(db: Session, event_id: UUID, content: bytes, filename: str, config, current):
    result = preview(db, event_id, content, filename, current)
    if result["errors"]:
        raise HTTPException(422, detail={"message": "Excel validation failed", **result})
    if result["file_sha256"] != config.file_sha256:
        raise HTTPException(409, "El archivo no coincide con el preview")
    version = db.get(LogbookTemplateVersion, config.template_version_id)
    if not version or version.status != LogbookVersionStatus.PUBLISHED or version.template.status == LogbookTemplateStatus.ARCHIVED:
        raise HTTPException(422, "Se requiere una versión publicada")
    staff = set(db.scalars(select(EventStaff.user_id).join(User).where(EventStaff.event_id == event_id, User.is_active.is_(True))).all())
    if not set(config.participant_ids) <= staff:
        raise HTTPException(422, "Todos los participantes deben pertenecer al evento")
    if config.supervisor_id:
        supervisor = db.scalar(select(User).join(EventStaff, EventStaff.user_id == User.id).where(EventStaff.event_id == event_id, User.id == config.supervisor_id, User.role == UserRole.SUPERVISOR, User.is_active.is_(True)))
        if not supervisor:
            raise HTTPException(422, "Supervisor inválido")
    try:
        zone = ZoneInfo(config.timezone)
    except ZoneInfoNotFoundError as exc:
        raise HTTPException(422, "Timezone inválida") from exc
    batch = LogbookImportBatch(event_id=event_id, original_filename=result["filename"], file_sha256=result["file_sha256"], sheet_name=result["sheet_name"], activities_count=result["activities_count"], dates_count=result["dates_count"], scheduled_items_count=result["scheduled_items_count"], instances_created=result["instances_to_create"], imported_by=current.id, configuration={"timezone": config.timezone, "participant_count": len(config.participant_ids), "client_visibility": config.client_visibility})
    db.add(batch)
    try:
        db.flush()
        created = []
        for day in result["days"]:
            if not day["activities"]:
                continue
            operational_date = date.fromisoformat(day["date"])
            opens_at = datetime.combine(operational_date, config.opens_at_local, zone).astimezone(timezone.utc)
            due_at = datetime.combine(operational_date, config.due_at_local, zone).astimezone(timezone.utc)
            if due_at <= opens_at:
                raise HTTPException(422, "La hora de vencimiento debe ser posterior a la apertura")
            instance = LogbookInstance(event_id=event_id, template_id=version.template_id, template_version_id=version.id, name=f"{config.base_name} · {operational_date:%d/%m/%Y}", operational_stage=version.template.operational_stage, assignment_mode=LogbookAssignmentMode.SHARED, opens_at=opens_at, due_at=due_at, supervisor_id=config.supervisor_id, status=LogbookInstanceStatus.SCHEDULED if opens_at > datetime.now(timezone.utc) else LogbookInstanceStatus.OPEN, client_visibility=config.client_visibility, created_by=current.id, occurrence_date=operational_date, import_batch_id=batch.id)
            db.add(instance)
            db.flush()
            for uid in config.participant_ids:
                db.add(LogbookAssignment(logbook_instance_id=instance.id, user_id=uid))
            for position, item in enumerate(day["activities"]):
                db.add(LogbookInstanceItem(instance_id=instance.id, title=item["title"], source_row=item["source_row"], position=position))
            created.append(instance.id)
        audit(db, current, "LOGBOOK_XLSX_IMPORTED", "LogbookImportBatch", batch.id, event_id=event_id, new={"instances_created": len(created), "file_sha256": result["file_sha256"]})
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(409, "Este archivo ya fue importado para el evento") from exc
    except Exception:
        db.rollback()
        raise
    return {"batch_id": batch.id, "instances_created": len(created), "instance_ids": created}


def _bulk_instances(db: Session, batch_id: UUID, payload, current):
    batch = db.get(LogbookImportBatch, batch_id)
    if not batch:
        raise HTTPException(404, "Lote de importación no encontrado")
    if not can_manage_event(current, batch.event_id, db):
        raise HTTPException(403, "Insufficient role")
    query = select(LogbookInstance).where(LogbookInstance.import_batch_id == batch.id)
    if payload.scope == "FUTURE":
        query = query.where(LogbookInstance.status == LogbookInstanceStatus.SCHEDULED)
    elif payload.scope == "DATES":
        query = query.where(LogbookInstance.occurrence_date.in_(payload.dates))
    instances = list(db.scalars(query.order_by(LogbookInstance.occurrence_date).with_for_update()).all())
    if not instances:
        raise HTTPException(422, "No hay bitácoras que coincidan con el alcance seleccionado")
    return batch, instances


def bulk_participants(db: Session, batch_id: UUID, payload, current, *, apply: bool):
    batch, instances = _bulk_instances(db, batch_id, payload, current)
    selected = set(payload.participant_ids)
    valid_staff = set(db.scalars(
        select(EventStaff.user_id).join(User).where(
            EventStaff.event_id == batch.event_id,
            User.is_active.is_(True),
            User.role.in_([UserRole.WORKER, UserRole.LOGISTICS_OPERATOR, UserRole.SUPERVISOR]),
        )
    ).all())
    if not selected <= valid_staff:
        raise HTTPException(422, "Todos los participantes deben ser personal operativo activo del evento")
    instance_ids = [instance.id for instance in instances]
    assignments = list(db.scalars(select(LogbookAssignment).where(
        LogbookAssignment.logbook_instance_id.in_(instance_ids)
    ).with_for_update()).all())
    by_instance = {}
    for assignment in assignments:
        by_instance.setdefault(assignment.logbook_instance_id, {})[assignment.user_id] = assignment
    contribution_assignments = set(db.scalars(select(LogbookItemContribution.assignment_id).where(
        LogbookItemContribution.instance_id.in_(instance_ids)
    )).all())
    additions = removals = preserved = historical = 0
    changes = []
    for instance in instances:
        existing = by_instance.get(instance.id, {})
        active = {uid for uid, assignment in existing.items()
                  if assignment.status != LogbookAssignmentStatus.CANCELLED}
        desired = (active | selected if payload.operation == "ADD" else
                   active - selected if payload.operation == "REMOVE" else selected)
        if not desired:
            raise HTTPException(422, f"La bitácora {instance.name} quedaría sin participantes")
        add_ids, remove_ids = desired - active, active - desired
        additions += len(add_ids)
        removals += len(remove_ids)
        preserved += len(active & desired)
        for uid in add_ids:
            changes.append(("add", instance, existing.get(uid), uid))
        for uid in remove_ids:
            assignment = existing[uid]
            has_history = assignment.id in contribution_assignments or bool(assignment.responses)
            historical += int(has_history)
            changes.append(("remove", instance, assignment, has_history))
    result = {
        "batch_id": batch.id, "operation": payload.operation, "scope": payload.scope,
        "instances_matched": len(instances), "assignments_to_add": additions,
        "assignments_to_remove": removals, "assignments_preserved": preserved,
        "historical_assignments_preserved": historical,
        "participant_ids": payload.participant_ids, "applied": apply,
    }
    if not apply:
        return result
    try:
        for action, instance, assignment, value in changes:
            if action == "add":
                if assignment:
                    assignment.status = (LogbookAssignmentStatus.IN_PROGRESS
                                         if assignment.id in contribution_assignments or assignment.responses
                                         else LogbookAssignmentStatus.PENDING)
                else:
                    db.add(LogbookAssignment(logbook_instance_id=instance.id, user_id=value))
            elif value:
                assignment.status = LogbookAssignmentStatus.CANCELLED
            else:
                db.delete(assignment)
        configuration = dict(batch.configuration or {})
        configuration["participant_count"] = db.scalar(select(func.count(func.distinct(LogbookAssignment.user_id))).join(
            LogbookInstance, LogbookInstance.id == LogbookAssignment.logbook_instance_id
        ).where(LogbookInstance.import_batch_id == batch.id,
                LogbookAssignment.status != LogbookAssignmentStatus.CANCELLED)) or 0
        batch.configuration = configuration
        audit(db, current, "LOGBOOK_IMPORT_PARTICIPANTS_BULK_UPDATED", "LogbookImportBatch",
              batch.id, event_id=batch.event_id, new=result)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return result


def bulk_supervisor(db: Session, batch_id: UUID, payload, current, *, apply: bool):
    batch, instances = _bulk_instances(db, batch_id, payload, current)
    if payload.supervisor_id:
        valid = db.scalar(select(EventStaff.user_id).join(User).where(
            EventStaff.event_id == batch.event_id,
            EventStaff.user_id == payload.supervisor_id,
            User.is_active.is_(True),
            User.role == UserRole.SUPERVISOR,
        ))
        if not valid:
            raise HTTPException(422, "El supervisor debe estar activo y asignado al evento")
    mutable_statuses = {
        LogbookInstanceStatus.DRAFT, LogbookInstanceStatus.SCHEDULED,
        LogbookInstanceStatus.OPEN, LogbookInstanceStatus.IN_PROGRESS,
        LogbookInstanceStatus.CHANGES_REQUESTED,
    }
    changeable = [item for item in instances if item.status in mutable_statuses]
    locked = len(instances) - len(changeable)
    changes = [item for item in changeable if item.supervisor_id != payload.supervisor_id]
    result = {
        "batch_id": batch.id, "scope": payload.scope,
        "supervisor_id": payload.supervisor_id,
        "instances_matched": len(instances), "instances_to_update": len(changes),
        "instances_unchanged": len(changeable) - len(changes),
        "instances_locked": locked, "applied": apply,
    }
    if not apply:
        return result
    for instance in changes:
        instance.supervisor_id = payload.supervisor_id
        instance.configuration_revision += 1
    configuration = dict(batch.configuration or {})
    configuration["supervisor_id"] = str(payload.supervisor_id) if payload.supervisor_id else None
    batch.configuration = configuration
    audit(db, current, "LOGBOOK_IMPORT_SUPERVISOR_BULK_UPDATED", "LogbookImportBatch",
          batch.id, event_id=batch.event_id, new=result)
    db.commit()
    return result
